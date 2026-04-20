from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.db.models import Count, Sum, Avg
from events.models import Event, Registration, Comment, Like, Favorite, Review, Category
from accounts.models import User
from datetime import datetime
import pandas as pd
from io import BytesIO
from django.db import models
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from events.models import OrganizerSubscription

def style_excel_sheet(ws, title, headers, filter_data=None):
    """
    Красивое оформление Excel-листа:
    - большой заголовок
    - строка с фильтрами
    - шапка таблицы
    - автоширина колонок
    """
    header_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
    title_fill = PatternFill(fill_type='solid', fgColor='B4C7E7')
    thin = Side(border_style='thin', color='000000')

    # Заголовок
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = title_fill
    ws['A1'].border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_row = 2

    # Фильтры
    if filter_data:
        for label, value in filter_data:
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=value if value else '—')
            current_row += 1

    # Пустая строка перед таблицей
    current_row += 1

    # Заголовки таблицы
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    return current_row + 1  # строка, с которой начинается таблица

@staff_member_required
def reports_dashboard(request):
    """Главная страница отчётов"""
    return render(request, 'reports/dashboard.html')

@staff_member_required
def organizer_report(request):
    """Отчёт по организаторам"""
    organizers = User.objects.filter(created_events__isnull=False).distinct().order_by('username')

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    organizer_id = request.GET.get('organizer')

    data = []
    filtered_total_events = 0

    for user in organizers:
        if organizer_id and str(user.id) != str(organizer_id):
            continue

        events = user.created_events.all()

        if date_from:
            events = events.filter(start_datetime__date__gte=date_from)
        if date_to:
            events = events.filter(start_datetime__date__lte=date_to)

        events_count = events.count()
        filtered_total_events += events_count

        total_participants = Registration.objects.filter(event__in=events, is_invited=False).count()
        total_likes = Like.objects.filter(event__in=events).count()

        if events_count > 0 or not any([date_from, date_to, organizer_id]):
            data.append({
                'user': user,
                'events_count': events_count,
                'total_participants': total_participants,
                'total_likes': total_likes,
            })

    data.sort(key=lambda x: x['events_count'], reverse=True)

    context = {
        'data': data,
        'total_organizers': len(data),
        'total_events': filtered_total_events,
        'date_from': date_from,
        'date_to': date_to,
        'selected_organizer': organizer_id,
        'organizers': organizers,
    }
    return render(request, 'reports/organizer_report.html', context)

@staff_member_required
def events_report(request):
    """Отчёт по мероприятиям"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    category_id = request.GET.get('category')
    status = request.GET.get('status')

    events = Event.objects.select_related('creator', 'category').all().order_by('-start_datetime')

    if date_from:
        events = events.filter(start_datetime__date__gte=date_from)
    if date_to:
        events = events.filter(start_datetime__date__lte=date_to)
    if category_id:
        events = events.filter(category_id=category_id)
    if status:
        events = events.filter(status=status)

    total_events = events.count()

    data = []
    for event in events:
        participants = Registration.objects.filter(event=event, is_invited=False).count()
        likes = Like.objects.filter(event=event).count()
        favorites = Favorite.objects.filter(event=event).count()

        data.append({
            'event': event,
            'participants': participants,
            'likes': likes,
            'favorites': favorites,
        })

    # Статистика по статусам
    status_stats = {}
    status_display = dict(Event.STATUS_CHOICES) if hasattr(Event, 'STATUS_CHOICES') else {}

    if total_events > 0:
        for key, label in status_display.items():
            count = events.filter(status=key).count()
            if count > 0:
                status_stats[label] = {
                    'count': count,
                    'percent': round((count / total_events) * 100, 1)
                }

    # Статистика по месяцам
    monthly_stats = []
    months = events.dates('start_datetime', 'month', order='DESC')
    for month in months[:6]:
        count = events.filter(
            start_datetime__year=month.year,
            start_datetime__month=month.month
        ).count()
        percent = round((count / total_events) * 100, 1) if total_events > 0 else 0
        monthly_stats.append({
            'month': month,
            'count': count,
            'percent': percent
        })

    context = {
        'data': data,
        'total_events': total_events,
        'status_stats': status_stats,
        'monthly_stats': monthly_stats,
        'categories': Category.objects.all(),
        'selected_category': category_id,
        'selected_status': status,
        'date_from': date_from,
        'date_to': date_to,
        'status_choices': Event.STATUS_CHOICES if hasattr(Event, 'STATUS_CHOICES') else [],
    }
    return render(request, 'reports/events_report.html', context)

@staff_member_required
def participants_report(request):
    """Отчёт по участникам"""
    participants = User.objects.filter(event_registrations__isnull=True).distinct()
    
    data = []
    for user in participants:
        registrations_count = user.event_registrations.count()
        events_attended = user.event_registrations.filter(status='attended').count()
        
        # Получаем категории мероприятий, которые посещал
        categories = []
        for reg in user.event_registrations.all():
            if reg.event.category:
                categories.append(reg.event.category.name)
        
        from collections import Counter
        top_categories = Counter(categories).most_common(3)
        
        data.append({
            'user': user,
            'registrations_count': registrations_count,
            'events_attended': events_attended,
            'top_categories': top_categories,
            'has_restrictions': user.dietary_restrictions.exists() if hasattr(user, 'dietary_restrictions') else False,
        })
    
    data.sort(key=lambda x: x['registrations_count'], reverse=True)
    
    context = {
        'data': data,
        'total_participants': participants.count(),
        'total_registrations': Registration.objects.count(),
    }
    return render(request, 'reports/participants_report.html', context)


@staff_member_required
def export_events_excel(request):
    """Экспорт отчёта по мероприятиям в Excel с учётом фильтров"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    category_id = request.GET.get('category')
    status = request.GET.get('status')

    events = Event.objects.select_related('creator', 'category').all().order_by('-start_datetime')

    if date_from:
        events = events.filter(start_datetime__date__gte=date_from)
    if date_to:
        events = events.filter(start_datetime__date__lte=date_to)
    if category_id:
        events = events.filter(category_id=category_id)
    if status:
        events = events.filter(status=status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Мероприятия"

    headers = [
        'Название',
        'Дата',
        'Организатор',
        'Категория',
        'Статус',
        'Участников',
        'Лайков',
        'В избранном'
    ]

    category_name = 'Все'
    if category_id:
        category = Category.objects.filter(id=category_id).first()
        if category:
            category_name = category.name

    status_name = status or 'Все'

    start_row = style_excel_sheet(
        ws,
        title='Отчёт по мероприятиям',
        headers=headers,
        filter_data=[
            ('Дата от', date_from or '—'),
            ('Дата до', date_to or '—'),
            ('Категория', category_name),
            ('Статус', status_name),
        ]
    )

    thin = Side(border_style='thin', color='000000')
    row_num = start_row

    for event in events:
        participants = Registration.objects.filter(event=event, is_invited=False).count()
        likes = Like.objects.filter(event=event).count()
        favorites = Favorite.objects.filter(event=event).count()

        values = [
            event.title,
            event.start_datetime.strftime('%d.%m.%Y %H:%M') if event.start_datetime else '',
            event.creator.username if event.creator else '',
            event.category.name if event.category else '—',
            event.get_status_display() if hasattr(event, 'get_status_display') else event.status,
            participants,
            likes,
            favorites,
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if col_num >= 6:
                cell.alignment = Alignment(horizontal='center')

        row_num += 1

    # Автоширина колонок
    for col_num in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=events_report.xlsx'
    wb.save(response)
    return response

@staff_member_required
def export_specialists_excel(request):
    """Экспорт специалистов в Excel"""
    specialists = User.objects.filter(created_events__isnull=False).distinct()
    
    df = pd.DataFrame([{
        'ID': u.id,
        'Имя пользователя': u.username,
        'Email': u.email,
        'Телефон': u.phone or '',
        'Мероприятий': u.created_events.count(),
        'Участников': Registration.objects.filter(event__creator=u).count(),
        'Лайков': Like.objects.filter(event__creator=u).count(),
        'Избранное': Favorite.objects.filter(event__creator=u).count(),
        'Средний рейтинг': round(Review.objects.filter(event__creator=u).aggregate(avg=models.Avg('rating'))['avg'] or 0, 1),
        'Навыки': u.skills or '',
        'Опыт': u.experience or '',
        'Дата регистрации': u.date_joined.strftime('%d.%m.%Y'),
    } for u in specialists])
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Специалисты', index=False)
    
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=specialists_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response

@staff_member_required
def export_organizers_excel(request):
    """Экспорт отчёта по организаторам в Excel с учётом фильтров"""
    organizers = User.objects.filter(created_events__isnull=False).distinct().order_by('username')

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    organizer_id = request.GET.get('organizer')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Организаторы"

    headers = [
        'Организатор',
        'Email',
        'Количество мероприятий',
        'Количество участников',
        'Количество лайков'
    ]

    organizer_name = 'Все'
    if organizer_id:
        selected = organizers.filter(id=organizer_id).first()
        if selected:
            organizer_name = selected.username

    start_row = style_excel_sheet(
        ws,
        title='Отчёт по организаторам',
        headers=headers,
        filter_data=[
            ('Дата от', date_from or '—'),
            ('Дата до', date_to or '—'),
            ('Организатор', organizer_name),
        ]
    )

    thin = Side(border_style='thin', color='000000')
    row_num = start_row

    for user in organizers:
        if organizer_id and str(user.id) != str(organizer_id):
            continue

        events = user.created_events.all()

        if date_from:
            events = events.filter(start_datetime__date__gte=date_from)
        if date_to:
            events = events.filter(start_datetime__date__lte=date_to)

        events_count = events.count()
        total_participants = Registration.objects.filter(event__in=events, is_invited=False).count()
        total_likes = Like.objects.filter(event__in=events).count()

        if events_count > 0 or not any([date_from, date_to, organizer_id]):
            values = [
                user.username,
                user.email,
                events_count,
                total_participants,
                total_likes
            ]
            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if col_num >= 3:
                    cell.alignment = Alignment(horizontal='center')
            row_num += 1

    # Автоширина колонок
    for col_num in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=organizers_report.xlsx'
    wb.save(response)
    return response

@login_required
def my_organizer_report(request):
    events = Event.objects.filter(creator=request.user).order_by('-start_datetime')

    total_events = events.count()
    published_events = events.filter(status='published').count()
    pending_events = events.filter(status='pending').count()
    rejected_events = events.filter(status='rejected').count()

    total_participants = Registration.objects.filter(
        event__in=events,
        status='confirmed',
        is_invited=False
    ).count()

    total_likes = Like.objects.filter(event__in=events).count()
    total_favorites = Favorite.objects.filter(event__in=events).count()
    total_comments = Comment.objects.filter(event__in=events).count()

    context = {
        'events': events[:10],
        'total_events': total_events,
        'published_events': published_events,
        'pending_events': pending_events,
        'rejected_events': rejected_events,
        'total_participants': total_participants,
        'total_likes': total_likes,
        'total_favorites': total_favorites,
        'total_comments': total_comments,
    }
    return render(request, 'reports/my_organizer_report.html', context)


@login_required
def my_participant_report(request):
    registrations = Registration.objects.filter(user=request.user).select_related('event').order_by('-created_at')
    upcoming_registrations = registrations.filter(event__start_datetime__gte=timezone.now())
    attended_count = registrations.filter(status='attended').count()

    favorite_events = Favorite.objects.filter(user=request.user).count()
    favorite_organizers = OrganizerSubscription.objects.filter(user=request.user).count()

    context = {
        'registrations': registrations[:10],
        'upcoming_registrations': upcoming_registrations[:5],
        'total_registrations': registrations.count(),
        'attended_count': attended_count,
        'favorite_events': favorite_events,
        'favorite_organizers': favorite_organizers,
    }
    return render(request, 'reports/my_participant_report.html', context)